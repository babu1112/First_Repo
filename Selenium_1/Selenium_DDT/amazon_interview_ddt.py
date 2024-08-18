from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from openpyxl import Workbook

# Define path to your Chrome webdriver (replace with actual path)
chromedriver_path = r"C:\Eclipse Workspace\chromedriver-win32\chromedriver.exe"
ser = Service(executable_path=chromedriver_path)

# Open a new workbook
wb = Workbook()
sheet = wb.active
sheet.cell(row=1, column=1).value = "Product Name"

# Launch Chrome browser
driver = webdriver.Chrome(service=ser)

# Open Amazon India
driver.get("https://www.amazon.in/")

# Search for "apple products"
search_box = driver.find_element(By.ID, "twotabsearchtextbox")
search_box.send_keys("apple products")
search_box.submit()

# Find product elements
#product_elements = driver.find_elements(By.XPATH, "//div[contains(@class,'s-product-grid')]")
#product_elements = driver.find_elements(By.XPATH, "//span[@class='a-size-medium a-color-base a-text-normal']")
#product_elements = driver.find_elements(By.XPATH, "//div//h2")
product_elements = driver.find_elements(By.XPATH,'//div[contains(@class,"a-section a-spacing-none puis-padding-right-small s-title-instructions-style")]')

# Extract names of first five products
for i in range(5):
    if i >= len(product_elements):
        break
    product_name = product_elements[i].find_element(By.TAG_NAME, "h2").text
    sheet.cell(row=i + 2, column=1).value = product_name

# Close the browser
driver.quit()

# Save the workbook
wb.save("apple_products.xlsx")  # Replace with your desired filename

print("Successfully extracted product names and saved to excel!")
